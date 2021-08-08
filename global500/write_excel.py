
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border,Side
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.utils import get_column_letter, column_index_from_string

from copy import copy
from get_data_global500 import get_data_global500
from get_industry_cat import get_industry_cat

def write_excel(industry_dict_2021, industry_dict_2020,  company_list_2021, company_list_2020, template_file='template.xlsx', out_file='result.xlsx'):

    font = Font(u'等线',size = 12, bold=True, italic=False, strike=False, color='000000')
    fill = PatternFill(fill_type='solid', start_color='E2EFDA', end_color='E2EFDA')
    border = Border(top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    template_wb = openpyxl.load_workbook(template_file)
    template_ws = template_wb["template"]

    wb = openpyxl.Workbook()
    er_dollar2cny = 7.0

    total_2020 = 0
    total_2021 = 0

    for key in industry_dict_2020:
        ws        = wb.create_sheet(str(industry_dict_2020[key]))
        ws.append(['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',''])
        ws.append([''])
        ws.append([''])

        wm = list(template_ws.merged_cells)   #开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        # print (list(wm))
        if len(wm)>0 :
            for i in range(0,len(wm)):
                cell2=str(wm[i]).replace('(<CellRange ','').replace('>,)','')
                #print("MergeCell : %s" % cell2)
                ws.merge_cells(cell2)

        for rows in range(3):
            ws.row_dimensions[rows+1].height = template_ws.row_dimensions[rows+1].height 
            for col in range(40):
                ws.column_dimensions[get_column_letter(col+1)].width = template_ws.column_dimensions[get_column_letter(col+1)].width
                ws.cell(row=rows+1,column=col+1,value=template_ws.cell(rows+1, col+1).value)

                if template_ws.cell(rows+1,  col+1).has_style:    #拷贝格式
                    ws.cell(row=rows+1,column=col+1).font = copy(template_ws.cell(rows+1, col+1).font)
                    ws.cell(row=rows+1,column=col+1).border = copy(template_ws.cell(rows+1, col+1).border)
                    ws.cell(row=rows+1,column=col+1).fill = copy(template_ws.cell(rows+1, col+1).fill)
                    ws.cell(row=rows+1,column=col+1).number_format = copy(template_ws.cell(rows+1, col+1).number_format)
                    ws.cell(row=rows+1,column=col+1).protection = copy(template_ws.cell(rows+1, col+1).protection)
                    ws.cell(row=rows+1,column=col+1).alignment = copy(template_ws.cell(rows+1, col+1).alignment)

        # menu_row  = ['2021年500强排名', '公司名', '销售额销售额\n百万美元', '国家', '员工数', '营收', '% vs 上一年', '利润', '% vs上一年', '总资产', '股东权益']
        # ws.append(menu_row)
        # ws['D2'] = '销售额M$'
        # ws['E2'] = '国家'
        # ws['F2'] = '员工数'
        # ws['G2'] = '营收'
        # ws['H2'] = '% vs 上一年'
        # ws['I2'] = '利润'
        # ws['J2'] = '% vs上一年'
        # ws['K2'] = '总资产'
        # ws['L2'] = '股东权益'

        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['C'].width = 32.0
        ws.column_dimensions['M'].width = 2
        ws.column_dimensions['S'].width = 32.0
        ws.column_dimensions['Q'].width = 4
        ws.column_dimensions['AC'].width = 2
        ws.column_dimensions['AG'].width = 3

        row = 4
        col = 1
        count_2020 = 0
        count_2021 = 0

        chinese_data = [[1,0], [1,0], [1,0]]
        non_chinese_data = [[1,0], [1,0], [1,0]]

        for company in company_list_2020:
            if company.industry == str(industry_dict_2020[key]):
                count_2020 += 1
                total_2020 += 1
                ws.append(['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',''])
                ws[row][col+0].value = company.rank
                ws[row][col+1].value = company.company_name
                ws[row][col+2].value = company.revenue
                ws[row][col+3].value = company.country
                ws[row][col+4].value = company.employee
                ws[row][col+5].value = company.revenue
                ws[row][col+6].value = company.revenue_change_percent
                ws[row][col+7].value = company.profit
                ws[row][col+8].value = company.profit_change_percent
                ws[row][col+9].value = company.total_money
                ws[row][col+10].value = company.shareholders_equity

                try:
                    ws[row][col+12].value = '%.1f%%' % (float(company.profit) / float(company.revenue) * 100)
                except:
                    ws[row][col+12].value = "--"
                try:
                    ws[row][col+13].value = '%.1f%%' % (float(company.profit) / float(company.shareholders_equity) * 100)
                except:
                    ws[row][col+13].value = "--"
                try:
                    ws[row][col+14].value = '%.1f' % (float(company.profit) / float(company.employee) * er_dollar2cny * 100)
                except:
                    ws[row][col+14].value = "--"

                ws[row][col+0].font = font
                ws[row][col+0].border=border
                ws[row][col+0].alignment=align
                ws[row][col+1].font = font
                ws[row][col+1].border=border
                ws[row][col+1].alignment=align
                ws[row][col+2].font = font
                ws[row][col+2].border=border
                ws[row][col+2].alignment=align
                ws[row][col+3].font = font
                ws[row][col+3].border=border
                ws[row][col+3].alignment=align
                ws[row][col+4].font = font
                ws[row][col+4].border=border
                ws[row][col+4].alignment=align
                ws[row][col+5].font = font
                ws[row][col+5].border=border
                ws[row][col+5].alignment=align
                ws[row][col+6].font = font
                ws[row][col+6].border=border
                ws[row][col+6].alignment=align
                ws[row][col+7].font = font
                ws[row][col+7].border=border
                ws[row][col+7].alignment=align
                ws[row][col+8].font = font
                ws[row][col+8].border=border
                ws[row][col+8].alignment=align
                ws[row][col+9].font = font
                ws[row][col+9].border=border
                ws[row][col+9].alignment=align
                ws[row][col+10].font = font
                ws[row][col+10].border=border
                ws[row][col+10].alignment=align
                ws[row][col+11].font = font
                ws[row][col+11].border=border
                ws[row][col+11].alignment=align
                ws[row][col+12].font = font
                ws[row][col+12].border=border
                ws[row][col+12].alignment=align
                ws[row][col+13].font = font
                ws[row][col+13].border=border
                ws[row][col+13].alignment=align
                ws[row][col+14].font = font
                ws[row][col+14].border=border
                ws[row][col+14].alignment=align

                if(company.country == "中国"):
                    ws[row][col+0].fill = fill
                    ws[row][col+1].fill = fill
                    ws[row][col+2].fill = fill
                    ws[row][col+3].fill = fill
                    ws[row][col+4].fill = fill
                    ws[row][col+5].fill = fill
                    ws[row][col+6].fill = fill
                    ws[row][col+7].fill = fill
                    ws[row][col+8].fill = fill
                    ws[row][col+9].fill = fill
                    ws[row][col+10].fill = fill
                    ws[row][col+11].fill = fill
                    ws[row][col+12].fill = fill
                    ws[row][col+13].fill = fill
                    ws[row][col+14].fill = fill

                ws.append([''])
                ws.row_dimensions[row+1].height = 6.5
                row = row + 2

                # data_first_line = [company.rank, 
                #                    company.company_name, 
                #                    company.revenue, 
                #                    company.country, 
                #                    company.employee, 
                #                    company.revenue, 
                #                    company.revenue_change_percent, 
                #                    company.profit, 
                #                    company.profit_change_percent, 
                #                    company.total_money, 
                #                    company.shareholders_equity]

        row = 4
        col = 17

        for company in company_list_2021:
            if company.industry == str(industry_dict_2020[key]):
                count_2021 += 1
                total_2021 += 1

                if(count_2021 > count_2020):
                    ws.append(['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',''])
                    ws.append([''])
                    ws.row_dimensions[row+1].height = 6.5

                ws[row][col+0].value = company.rank
                ws[row][col+1].value = company.company_name
                ws[row][col+2].value = company.revenue
                ws[row][col+3].value = company.country
                ws[row][col+4].value = company.employee
                ws[row][col+5].value = company.revenue
                ws[row][col+6].value = company.revenue_change_percent
                ws[row][col+7].value = company.profit
                ws[row][col+8].value = company.profit_change_percent
                ws[row][col+9].value = company.total_money
                ws[row][col+10].value = company.shareholders_equity

                try:
                    ws[row][col+12].value = '%.1f%%' % (float(company.profit) / float(company.revenue) * 100)
                except:
                    ws[row][col+12].value = "--"
                try:
                    ws[row][col+13].value = '%.1f%%' % (float(company.profit) / float(company.shareholders_equity) * 100)
                except:
                    ws[row][col+13].value = "--"
                try:
                    ws[row][col+14].value = '%.1f' % (float(company.profit) / float(company.employee) * er_dollar2cny * 100)
                except:
                    ws[row][col+14].value = "--"

                ws[row][col+0].font = font
                ws[row][col+0].border=border
                ws[row][col+0].alignment=align
                ws[row][col+1].font = font
                ws[row][col+1].border=border
                ws[row][col+1].alignment=align
                ws[row][col+2].font = font
                ws[row][col+2].border=border
                ws[row][col+2].alignment=align
                ws[row][col+3].font = font
                ws[row][col+3].border=border
                ws[row][col+3].alignment=align
                ws[row][col+4].font = font
                ws[row][col+4].border=border
                ws[row][col+4].alignment=align
                ws[row][col+5].font = font
                ws[row][col+5].border=border
                ws[row][col+5].alignment=align
                ws[row][col+6].font = font
                ws[row][col+6].border=border
                ws[row][col+6].alignment=align
                ws[row][col+7].font = font
                ws[row][col+7].border=border
                ws[row][col+7].alignment=align
                ws[row][col+8].font = font
                ws[row][col+8].border=border
                ws[row][col+8].alignment=align
                ws[row][col+9].font = font
                ws[row][col+9].border=border
                ws[row][col+9].alignment=align
                ws[row][col+10].font = font
                ws[row][col+10].border=border
                ws[row][col+10].alignment=align
                ws[row][col+11].font = font
                ws[row][col+11].border=border
                ws[row][col+11].alignment=align
                ws[row][col+12].font = font
                ws[row][col+12].border=border
                ws[row][col+12].alignment=align
                ws[row][col+13].font = font
                ws[row][col+13].border=border
                ws[row][col+13].alignment=align
                ws[row][col+14].font = font
                ws[row][col+14].border=border
                ws[row][col+14].alignment=align

                if(company.country == "中国"):
                    try:
                        chinese_data[0][1] += (float(company.profit) / float(company.revenue) * 100)
                        chinese_data[0][0] += 1
                    except:
                        pass
                    try:
                        chinese_data[1][1] += (float(company.profit) / float(company.shareholders_equity) * 100)
                        chinese_data[1][0] += 1
                    except:
                        pass
                    try:
                        chinese_data[2][1] += (float(company.profit) / float(company.employee) * er_dollar2cny * 100)
                        chinese_data[2][0] += 1
                    except:
                        pass 

                    ws[row][col+0].fill = fill
                    ws[row][col+1].fill = fill
                    ws[row][col+2].fill = fill
                    ws[row][col+3].fill = fill
                    ws[row][col+4].fill = fill
                    ws[row][col+5].fill = fill
                    ws[row][col+6].fill = fill
                    ws[row][col+7].fill = fill
                    ws[row][col+8].fill = fill
                    ws[row][col+9].fill = fill
                    ws[row][col+10].fill = fill
                    ws[row][col+11].fill = fill
                    ws[row][col+12].fill = fill
                    ws[row][col+13].fill = fill
                    ws[row][col+14].fill = fill

                else:
                    try:
                        non_chinese_data[0][1] += (float(company.profit) / float(company.revenue) * 100)
                        non_chinese_data[0][0] += 1
                    except:
                        pass
                    try:
                        non_chinese_data[1][1] += (float(company.profit) / float(company.shareholders_equity) * 100)
                        non_chinese_data[1][0] += 1
                    except:
                        pass
                    try:
                        non_chinese_data[2][1] += (float(company.profit) / float(company.employee) * er_dollar2cny * 100)
                        non_chinese_data[2][0] += 1
                    except:
                        pass
                
                row = row + 2

        try:
            ws[4][33].value =  non_chinese_data[0][1] / (non_chinese_data[0][0] - 1) / 100
            ws[4][33].number_format = '0.0%'
        except:
            ws[4][33].value = 0

        try:
            ws[4][35].value =  non_chinese_data[1][1] / (non_chinese_data[1][0] - 1) / 100
            ws[4][35].number_format = '0.0%'
        except:
            ws[4][35].value = 0

        try:
            ws[4][37].value = non_chinese_data[2][1]   / (non_chinese_data[2][0] - 1)
            ws[4][37].number_format = '0.0'
        except:
            ws[4][37].value = 0
        
        try:
            ws[4][34].value =  chinese_data[0][1]     / (chinese_data[0][0] - 1) / 100
            ws[4][34].number_format = '0.0%'
        except:
            ws[4][34].value = 0
        
        try:
            ws[4][36].value =  chinese_data[1][1]     / (chinese_data[1][0] - 1) / 100
            ws[4][36].number_format = '0.0%'
        except:
            ws[4][36].value = 0
        try:
            ws[4][38].value =  chinese_data[2][1]       / (chinese_data[2][0] - 1)
            ws[4][38].number_format = '0.0'
        except:
            ws[4][38].value = 0

        ws[4][33].font = font
        ws[4][33].border = border
        ws[4][33].alignment = align
        ws[4][35].font = font
        ws[4][35].border = border
        ws[4][35].alignment = align
        ws[4][37].font = font
        ws[4][37].border = border
        ws[4][37].alignment = align
        ws[4][34].font = font
        ws[4][34].border = border
        ws[4][34].alignment = align
        ws[4][36].font = font
        ws[4][36].border = border
        ws[4][36].alignment = align
        ws[4][38].font = font
        ws[4][38].border = border
        ws[4][38].alignment = align

        # Chart

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "平均利润率"
        chart1.y_axis.title = ''
        chart1.x_axis.title = ''

        data = Reference(ws, min_col=34, max_col=35, min_row=4, max_row=4)
        cats = Reference(ws, min_col=34, max_col=35, min_row=3, max_row=3)
        chart1.add_data(data)
        chart1.set_categories(cats)
        chart1.shape = 1
        ws.add_chart(chart1, "AH6")


    for key in industry_dict_2021:
        flag = False
        for key_2020 in industry_dict_2020:
            if industry_dict_2021[key] == industry_dict_2020[key_2020]:
                flag = True
                break

        if not flag:
            # print(str(industry_dict_2021[key]))
            ws        = wb.create_sheet(str(industry_dict_2021[key]))
            ws.append(['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',''])
            ws.append([''])
            ws.append([''])

            wm = list(template_ws.merged_cells)   #开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
            # print (list(wm))
            if len(wm)>0 :
                for i in range(0,len(wm)):
                    cell2=str(wm[i]).replace('(<CellRange ','').replace('>,)','')
                    #print("MergeCell : %s" % cell2)
                    ws.merge_cells(cell2)

            for rows in range(3):
                ws.row_dimensions[rows+1].height = template_ws.row_dimensions[rows+1].height 
                for col in range(40):
                    ws.column_dimensions[get_column_letter(col+1)].width = template_ws.column_dimensions[get_column_letter(col+1)].width
                    ws.cell(row=rows+1,column=col+1,value=template_ws.cell(rows+1, col+1).value)

                    if template_ws.cell(rows+1,  col+1).has_style:    #拷贝格式
                        ws.cell(row=rows+1,column=col+1).font = copy(template_ws.cell(rows+1, col+1).font)
                        ws.cell(row=rows+1,column=col+1).border = copy(template_ws.cell(rows+1, col+1).border)
                        ws.cell(row=rows+1,column=col+1).fill = copy(template_ws.cell(rows+1, col+1).fill)
                        ws.cell(row=rows+1,column=col+1).number_format = copy(template_ws.cell(rows+1, col+1).number_format)
                        ws.cell(row=rows+1,column=col+1).protection = copy(template_ws.cell(rows+1, col+1).protection)
                        ws.cell(row=rows+1,column=col+1).alignment = copy(template_ws.cell(rows+1, col+1).alignment)

            # menu_row  = ['2021年500强排名', '公司名', '销售额销售额\n百万美元', '国家', '员工数', '营收', '% vs 上一年', '利润', '% vs上一年', '总资产', '股东权益']
            # ws.append(menu_row)
            # ws['D2'] = '销售额M$'
            # ws['E2'] = '国家'
            # ws['F2'] = '员工数'
            # ws['G2'] = '营收'
            # ws['H2'] = '% vs 上一年'
            # ws['I2'] = '利润'
            # ws['J2'] = '% vs上一年'
            # ws['K2'] = '总资产'
            # ws['L2'] = '股东权益'

            ws.column_dimensions['A'].width = 3
            ws.column_dimensions['C'].width = 32.0
            ws.column_dimensions['M'].width = 2
            ws.column_dimensions['S'].width = 32.0
            ws.column_dimensions['Q'].width = 4
            ws.column_dimensions['AC'].width = 2
            ws.column_dimensions['AG'].width = 3
    
            row = 4
            col = 17

            for company in company_list_2021:
                if company.industry == str(industry_dict_2021[key]):
                    count_2021 += 1
                    total_2021 += 1

                    ws.append(['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',''])
                    ws.append([''])
                    ws.row_dimensions[row+1].height = 6.5

                    ws[row][col+0].value = company.rank
                    ws[row][col+1].value = company.company_name
                    ws[row][col+2].value = company.revenue
                    ws[row][col+3].value = company.country
                    ws[row][col+4].value = company.employee
                    ws[row][col+5].value = company.revenue
                    ws[row][col+6].value = company.revenue_change_percent
                    ws[row][col+7].value = company.profit
                    ws[row][col+8].value = company.profit_change_percent
                    ws[row][col+9].value = company.total_money
                    ws[row][col+10].value = company.shareholders_equity

                    try:
                        ws[row][col+12].value = '%.1f%%' % (float(company.profit) / float(company.revenue) * 100)
                    except:
                        ws[row][col+12].value = "--"
                    try:
                        ws[row][col+13].value = '%.1f%%' % (float(company.profit) / float(company.shareholders_equity) * 100)
                    except:
                        ws[row][col+13].value = "--"
                    try:
                        ws[row][col+14].value = '%.1f' % (float(company.profit) / float(company.employee) * er_dollar2cny * 100)
                    except:
                        ws[row][col+14].value = "--"

                    ws[row][col+0].font = font
                    ws[row][col+0].border=border
                    ws[row][col+0].alignment=align
                    ws[row][col+1].font = font
                    ws[row][col+1].border=border
                    ws[row][col+1].alignment=align
                    ws[row][col+2].font = font
                    ws[row][col+2].border=border
                    ws[row][col+2].alignment=align
                    ws[row][col+3].font = font
                    ws[row][col+3].border=border
                    ws[row][col+3].alignment=align
                    ws[row][col+4].font = font
                    ws[row][col+4].border=border
                    ws[row][col+4].alignment=align
                    ws[row][col+5].font = font
                    ws[row][col+5].border=border
                    ws[row][col+5].alignment=align
                    ws[row][col+6].font = font
                    ws[row][col+6].border=border
                    ws[row][col+6].alignment=align
                    ws[row][col+7].font = font
                    ws[row][col+7].border=border
                    ws[row][col+7].alignment=align
                    ws[row][col+8].font = font
                    ws[row][col+8].border=border
                    ws[row][col+8].alignment=align
                    ws[row][col+9].font = font
                    ws[row][col+9].border=border
                    ws[row][col+9].alignment=align
                    ws[row][col+10].font = font
                    ws[row][col+10].border=border
                    ws[row][col+10].alignment=align
                    ws[row][col+11].font = font
                    ws[row][col+11].border=border
                    ws[row][col+11].alignment=align
                    ws[row][col+12].font = font
                    ws[row][col+12].border=border
                    ws[row][col+12].alignment=align
                    ws[row][col+13].font = font
                    ws[row][col+13].border=border
                    ws[row][col+13].alignment=align
                    ws[row][col+14].font = font
                    ws[row][col+14].border=border
                    ws[row][col+14].alignment=align

                    if(company.country == "中国"):
                        try:
                            chinese_data[0][1] += (float(company.profit) / float(company.revenue) * 100)
                            chinese_data[0][0] += 1
                        except:
                            pass
                        try:
                            chinese_data[1][1] += (float(company.profit) / float(company.shareholders_equity) * 100)
                            chinese_data[1][0] += 1
                        except:
                            pass
                        try:
                            chinese_data[2][1] += (float(company.profit) / float(company.employee) * er_dollar2cny * 100)
                            chinese_data[2][0] += 1
                        except:
                            pass 

                        ws[row][col+0].fill = fill
                        ws[row][col+1].fill = fill
                        ws[row][col+2].fill = fill
                        ws[row][col+3].fill = fill
                        ws[row][col+4].fill = fill
                        ws[row][col+5].fill = fill
                        ws[row][col+6].fill = fill
                        ws[row][col+7].fill = fill
                        ws[row][col+8].fill = fill
                        ws[row][col+9].fill = fill
                        ws[row][col+10].fill = fill
                        ws[row][col+11].fill = fill
                        ws[row][col+12].fill = fill
                        ws[row][col+13].fill = fill
                        ws[row][col+14].fill = fill

                    else:
                        try:
                            non_chinese_data[0][1] += (float(company.profit) / float(company.revenue) * 100)
                            non_chinese_data[0][0] += 1
                        except:
                            pass
                        try:
                            non_chinese_data[1][1] += (float(company.profit) / float(company.shareholders_equity) * 100)
                            non_chinese_data[1][0] += 1
                        except:
                            pass
                        try:
                            non_chinese_data[2][1] += (float(company.profit) / float(company.employee) * er_dollar2cny * 100)
                            non_chinese_data[2][0] += 1
                        except:
                            pass
                    
                    row = row + 2

            try:
                ws[4][33].value =  non_chinese_data[0][1] / (non_chinese_data[0][0] - 1) / 100
                ws[4][33].number_format = '0.0%'
            except:
                ws[4][33].value = 0

            try:
                ws[4][35].value =  non_chinese_data[1][1] / (non_chinese_data[1][0] - 1) / 100
                ws[4][35].number_format = '0.0%'
            except:
                ws[4][35].value = 0

            try:
                ws[4][37].value = non_chinese_data[2][1]   / (non_chinese_data[2][0] - 1)
                ws[4][37].number_format = '0.0'
            except:
                ws[4][37].value = 0
            
            try:
                ws[4][34].value =  chinese_data[0][1]     / (chinese_data[0][0] - 1) / 100
                ws[4][34].number_format = '0.0%'
            except:
                ws[4][34].value = 0
            
            try:
                ws[4][36].value =  chinese_data[1][1]     / (chinese_data[1][0] - 1) / 100
                ws[4][36].number_format = '0.0%'
            except:
                ws[4][36].value = 0
            try:
                ws[4][38].value =  chinese_data[2][1]       / (chinese_data[2][0] - 1)
                ws[4][38].number_format = '0.0'
            except:
                ws[4][38].value = 0

            ws[4][33].font = font
            ws[4][33].border = border
            ws[4][33].alignment = align
            ws[4][35].font = font
            ws[4][35].border = border
            ws[4][35].alignment = align
            ws[4][37].font = font
            ws[4][37].border = border
            ws[4][37].alignment = align
            ws[4][34].font = font
            ws[4][34].border = border
            ws[4][34].alignment = align
            ws[4][36].font = font
            ws[4][36].border = border
            ws[4][36].alignment = align
            ws[4][38].font = font
            ws[4][38].border = border
            ws[4][38].alignment = align

            # Chart

            chart1 = BarChart()
            chart1.type = "col"
            chart1.style = 10
            chart1.title = "平均利润率"
            chart1.y_axis.title = ''
            chart1.x_axis.title = ''

            data = Reference(ws, min_col=34, max_col=35, min_row=4, max_row=4)
            cats = Reference(ws, min_col=34, max_col=35, min_row=3, max_row=3)
            chart1.add_data(data)
            chart1.set_categories(cats)
            chart1.shape = 1
            ws.add_chart(chart1, "AH6")

    wb.save(out_file)

    print(total_2020, total_2021)

    print("[+] Info: Write Excel Done!")

if __name__ == '__main__':
    root_url_2021       = 'https://www.caifuzhongwen.com/fortune500/paiming/global500/2021_%e4%b8%96%e7%95%8c500%e5%bc%ba.htm'
    root_url_2020       = 'https://www.caifuzhongwen.com/fortune500/paiming/global500/2020_%e4%b8%96%e7%95%8c500%e5%bc%ba.htm'
    industry_dict_2021  = get_industry_cat(root_url_2021,  cat_file='cat2021.json')
    industry_dict_2020  = get_industry_cat(root_url_2020,  cat_file='cat2020.json')
    company_list_2021   = get_data_global500(root_url_2021,  data_json="data2021.json", url_tag=12)
    company_list_2020   = get_data_global500(root_url_2020,  data_json="data2020.json", url_tag=10)
    write_excel(industry_dict_2021, industry_dict_2020,  company_list_2021, company_list_2020)

