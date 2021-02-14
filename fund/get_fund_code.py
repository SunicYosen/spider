'''
 Get Fund Array
'''

import re
from openpyxl import load_workbook

def get_fund_code(excel_file='data.xlsx'):
    code_re         = r'^\d{6}$'
    fund_code_array = []

    wb          = load_workbook(excel_file)
    name_list   = wb.sheetnames
    my_sheet    = wb[name_list[0]]

    code_row    = list(my_sheet.columns)[1]

    for code in code_row:
        if code.value == None:
            continue

        code_matched = re.findall(code_re, code.value)
        if code_matched:
            fund_code_array.append(code_matched[0])

    return fund_code_array

if __name__ == '__main__':
    excel_file = 'data.xlsx'
    fund_code_array = get_fund_code(excel_file)
    print(fund_code_array)