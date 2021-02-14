'''
 Write To Excel
'''

import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def write_excel(data_dict={}, excel_file='data.xlsx'):
    code_re     = r'^\d{6}$'
    wb          = load_workbook(excel_file)

    name_list   = wb.sheetnames
    my_sheet    = wb[name_list[0]]

    for row in my_sheet.rows:
        fund_code    = list(row)[1].value

        if fund_code == None:
            continue

        code_matched = re.findall(code_re, fund_code)

        if code_matched:
            try:
                fund_data          = data_dict[fund_code]
            except:
                fund_data          = []
                print('[-] Warning: {} No Data!')
                continue

            fund_data_array    = fund_data.get_data_array()
            fund_data_array[1] = fund_code

            index = 0
            for row_cell in list(row):
                row_cell.alignment = Alignment(horizontal='general',
                                               vertical='center',
                                               text_rotation=0,
                                               wrap_text=True,
                                               shrink_to_fit=False,
                                               indent=0)

                row_cell.value = fund_data_array[index]
                index += 1

    wb.save(excel_file)

if __name__ == '__main__':
    write_excel()
