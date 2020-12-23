
import openpyxl
from load_json import load_json
from set_data import set_data_from_arr

def write_excel(cities_arr, province_dict, out_file='result.xlsx'):
    wb = openpyxl.Workbook()
    sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']

    ws0       = wb.create_sheet(sheets[0])
    ws1       = wb.create_sheet(sheets[1])
    ws2       = wb.create_sheet(sheets[2])
    ws3       = wb.create_sheet(sheets[3])
    # Write Sheet1
    menu0_row  = ['省份名', '点位类型', '数量', '数量占比']
    menu1_row  = ['城市名', '城市类别', '点位类型', '数量', '数量占比']
    menu2_row  = ['省份名', '点位类型', '数量', '数量占比']
    menu3_row  = ['城市名', '城市类别', '点位场景', '数量', '数量占比']
    ws0.append(menu0_row)
    ws1.append(menu1_row)
    ws2.append(menu2_row)
    ws3.append(menu3_row)

    for name in province_dict:
        province   = province_dict[name]
        total_type_points = province.total_type_points
        total_scene_points = province.total_scene_points
        for device_type in province.province_device_types_distribution:
            type_nums = province.province_device_types_distribution[device_type]['点位数量']
            row_item  = [name, device_type, type_nums, type_nums/total_type_points]
            ws0.append(row_item)

        for device_scene in province.province_device_scenes_distribution:
            scene_nums = province.province_device_scenes_distribution[device_scene]['点位数量']
            row_item = [name, device_scene, scene_nums, scene_nums/total_scene_points]
            ws2.append(row_item)

    for city in cities_arr:

        city_name  = city.city_name
        city_rank  = city.city_rank
        
        total_type_points  = city.total_type_points
        total_scene_points = city.total_scene_points

        for device_type in city.city_device_types_distribution:
            type_nums = city.city_device_types_distribution[device_type]['点位数量']
            row_item  = [city_name, city_rank, device_type, type_nums, type_nums/total_type_points]
            ws1.append(row_item)

        for device_scene in city.city_device_scenes_distribution:
            scene_nums = city.city_device_scenes_distribution[device_scene]['点位数量']
            row_item   = [city_name, city_rank, device_scene, type_nums, scene_nums/total_scene_points]
            ws3.append(row_item)

    wb.save(out_file)


if __name__ == '__main__':
    json_file   = 'data.json'
    out_file    = 'result.xlsx'
    cities_data = load_json(json_file)
    cities_arr, province_dict = set_data_from_arr(cities_data)
    write_excel(cities_arr, province_dict, out_file)
